"""客服題庫批次自動化腳本。

把客服測試題庫中「尚未作答」的題目（col5 有題目、col8 為空），
依序自動送進 addwii Streamlit Bot 的「產品知識問答」分頁，
擷取 AI 回答後寫回 Excel 的 col8。

設計重點（皆為 Session 2 實測踩坑後的修正）：
- 用 Playwright 同步 API + Chromium（headless 預設），openpyxl 讀寫 Excel。
- Streamlit text_input 是 server-side session_state，JS 注入 value 後必須等 ~3s
  讓 websocket 把值 commit 到後端，否則點送出會送出空字串。
- 送出後 Streamlit 重渲染常跳回第一個 tab，需重點「產品知識問答」tab 一次。
- 答案容器會殘留上一題舊答案，因此完成判斷以「無 spinner + 答案文字不同於送出前
  + 長度 > 15」三條件同時成立為準，避免抓到舊答案。
- 每題即 wb.save()，確保中途崩潰不丟進度（冪等、可中斷續跑）。

用法：
    PYTHONUTF8=1 python tools/batch_faq_query.py --dry-run
    PYTHONUTF8=1 python tools/batch_faq_query.py --limit 50
    PYTHONUTF8=1 python tools/batch_faq_query.py --headed --bot-url http://192.168.23.140:8505/
"""
import argparse
import os
import sys
import time
from pathlib import Path

# Playwright / openpyxl 為外部依賴，啟動時就明確檢查，避免跑到一半才爆。
try:
    from playwright.sync_api import (
        Error as PlaywrightError,
        sync_playwright,
    )
except ImportError:
    print(
        "錯誤：缺少 playwright 套件。請執行：\n"
        "    python -m pip install playwright\n"
        "    python -m playwright install chromium",
        file=sys.stderr,
    )
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "錯誤：缺少 openpyxl 套件。請執行：python -m pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


# ---- 常數設定 ----
DEFAULT_BOT_URL = "http://192.168.23.140:8505/"
EXCEL_NAME = "空氣清淨機_客服Bot_測試題庫_4題.xlsx"
SHEET_NAME = "測試題庫"
QUESTION_COL = 5  # col5：問題來源欄
ANSWER_COL = 8    # col8：AI 回答寫入欄

ANSWER_MARKER = "💬 AI 回答"          # 答案區塊的標題文字
MIN_ANSWER_LEN = 8                    # 答案長度門檻；放寬到 8 避免合法短答（如「可以，支援貨到付款。」）誤判逾時
ANSWER_POLL_TIMEOUT = 45              # 單題輪詢上限（秒）
POLL_INTERVAL = 1.5                   # 每次輪詢間隔（秒）
COMMIT_WAIT = 1.0                     # fill() 後等 websocket 把值 commit 的短等待
RETAB_POLL_ITERATION = 3              # 輪詢到第幾次時補點一次 tab（約第 4-5 秒）
MAX_CONSECUTIVE_TIMEOUT = 3           # 連續 N 題 TIMEOUT 視為後端掛掉
THROTTLE_BETWEEN_QUESTIONS = 1.5      # 題與題之間的節流間隔（秒）

TIMEOUT_MARK = "__TIMEOUT__"          # 寫入 Excel 的逾時標記
BACKEND_ERROR_KEYWORDS = ("Traceback", "ModuleNotFoundError", "TypeError")

# 等「可見輸入框」出現的條件式（產品知識問答分頁的問題輸入框 placeholder 含 CADR/坪）。
# 不用 page.evaluate 合成 input 事件注入問題：Streamlit text_input 不認 isTrusted=false
# 的合成事件，值不會 commit 到 server-side session_state，後端拿空問題空跑不答。
# 改用 Playwright 原生 inp.fill()（送真實 input 事件，Streamlit 會 commit）。
WAIT_VISIBLE_INPUT_JS = (
    "() => [...document.querySelectorAll('input[type=text]')]"
    ".some(i => i.offsetParent !== null && /CADR|坪/.test(i.placeholder||''))"
)

# 擷取答案用的 JS：找含 marker 的 stMarkdownContainer，取其「後一個」元素的文字。
# 先用 offsetParent 過濾掉隱形元素，避免抓到其他 tab 的隱形重複 marker 殘影。
EXTRACT_ANSWER_JS = r"""
(marker) => {
  const els = Array.from(
    document.querySelectorAll('[data-testid="stMarkdownContainer"]'))
    .filter(el => el.offsetParent !== null);  // 只保留可見容器
  const idx = els.findIndex(el => el.innerText.includes(marker));
  if (idx >= 0 && els[idx + 1]) return els[idx + 1].innerText.trim();
  return "";
}
"""

# 偵測後端例外：Streamlit 把 Python Traceback 渲染在 stException 容器。
DETECT_EXCEPTION_JS = r"""
() => {
  const el = document.querySelector('[data-testid="stException"]');
  return el ? el.innerText.trim() : "";
}
"""

# 偵測 spinner 是否還在（答案尚未生成完成）。
HAS_SPINNER_JS = r"""
() => !!document.querySelector('.stSpinner, [data-testid="stSpinner"]')
"""


class BackendError(Exception):
    """後端真的吐出 Python 例外（Traceback/ModuleNotFoundError/TypeError）時拋出。

    與渲染類暫態 RuntimeError 區分：BackendError 代表後端答案管線壞掉、硬跑無意義，
    必須中止整批；渲染類 RuntimeError（reload 後 tab/input 尚未 render）只是單題暫態，
    降級為該題失敗續跑，絕不可殺整批。
    """


def get_excel_path() -> Path:
    """回傳題庫 Excel 的可攜式路徑（不硬編碼使用者名稱）。"""
    return Path.home() / "Downloads" / EXCEL_NAME


def find_resume_rows(ws, limit=None):
    """掃描工作表，回傳待處理的 (row_idx, question) 清單。

    條件：col5 有非空題目，且 col8 為空 → 納入。已答的列一律跳過，
    達成冪等與中斷續跑。limit 限制最多取幾題。
    """
    pending = []
    for row_idx in range(1, ws.max_row + 1):
        question = ws.cell(row=row_idx, column=QUESTION_COL).value
        answer = ws.cell(row=row_idx, column=ANSWER_COL).value
        if question is None or str(question).strip() == "":
            continue  # 空題跳過
        if answer is not None and str(answer).strip() != "":
            continue  # 已答跳過
        pending.append((row_idx, str(question).strip()))
        if limit is not None and len(pending) >= limit:
            break
    return pending


def wait_streamlit_ready(page):
    """等 Streamlit 載入完成：tab 元素出現代表頁面已 render。"""
    page.wait_for_selector('[role="tab"]', timeout=30000)


def click_tab_by_text(page, text):
    """點含特定文字的 tab。找不到時拋 RuntimeError 由呼叫端決定處理。"""
    tabs = page.query_selector_all('[role="tab"]')
    for tab in tabs:
        if text in (tab.inner_text() or ""):
            tab.click()
            return
    raise RuntimeError(f"找不到含「{text}」的 tab")


def click_button_retry(page, text, tries=8):
    """重試版點按鈕：fill 後 Streamlit 可能短暫重渲染把按鈕洗掉，重試確保點到。

    只點可見按鈕（避開其他 tab 隱形同名鈕），成功回傳 True、全數失敗回傳 False。
    """
    for _ in range(tries):
        for btn in page.query_selector_all("button"):
            try:
                if btn.is_visible() and text in (btn.inner_text() or ""):
                    btn.click()
                    return True
            except PlaywrightError:
                pass  # 按鈕重渲染中被 detach，下一輪重試
        time.sleep(1)
    return False


def get_question_input(page):
    """取得產品知識問答分頁的可見問題輸入框（placeholder 含 CADR/坪），無則回 None。"""
    for inp in page.query_selector_all("input[type=text]"):
        placeholder = inp.get_attribute("placeholder") or ""
        if inp.is_visible() and ("CADR" in placeholder or "坪" in placeholder):
            return inp
    return None


def check_backend_exception(page):
    """偵測後端 Python 例外，命中關鍵字則回傳例外文字，否則回傳空字串。"""
    exc_text = page.evaluate(DETECT_EXCEPTION_JS)
    if exc_text and any(k in exc_text for k in BACKEND_ERROR_KEYWORDS):
        return exc_text
    return ""


def extract_answer(page, prev_answer):
    """輪詢擷取新答案。

    完成判斷三條件同時成立：無 spinner、答案 != prev_answer、長度 > 門檻。
    逾時回傳 None（由呼叫端標記 TIMEOUT）。prev_answer 比對是為了避開
    答案容器殘留的上一題舊答案（實測踩坑）。
    """
    deadline = time.monotonic() + ANSWER_POLL_TIMEOUT
    iteration = 0
    while time.monotonic() < deadline:
        iteration += 1
        # 答案算完時（送出後約 5-8s）Streamlit 會跳回第一個 tab 把答案藏起來，
        # 故在輪詢途中補點一次 tab 把產品知識問答區重新顯示出來再擷取。
        if iteration == RETAB_POLL_ITERATION:
            try:
                click_tab_by_text(page, "產品知識問答")
            except RuntimeError:
                pass  # tab 暫時不在也不中斷輪詢
        if page.evaluate(HAS_SPINNER_JS):
            time.sleep(POLL_INTERVAL)
            continue
        current = page.evaluate(EXTRACT_ANSWER_JS, ANSWER_MARKER)
        if current and current != prev_answer and len(current) > MIN_ANSWER_LEN:
            return current
        time.sleep(POLL_INTERVAL)
    return None


def ask_one_question(page, question, bot_url):
    """送出單一問題並回傳答案文字；逾時回傳 None。

    實證配方：reload 開全新 session→選 tab→等可見 input→取 input→fill→
    等 commit→重試送出→輪詢擷取（輪詢途中補點一次 tab）。
    每一步的等待都對應一個實測踩坑點。
    """
    # 1. 每題前 reload 開全新 session。關鍵：Qwen 對話記憶存在後端 session_state，
    #    「清除」鈕只清畫面答案、不清後端對話 context，同一 session 連續問會串味
    #    （實證：Row 90 在批次裡被 Row 89 發票主題污染）。reload 開全新 session 才
    #    能讓各題獨立不互相污染。代價是每題多一次 reload（~4-5s），315 題可接受。
    page.goto(bot_url, timeout=60000)
    wait_streamlit_ready(page)

    # 2. 切到產品知識問答分頁，等「可見輸入框」出現（以 input 為準而非清除鈕，
    #    因清除鈕出現時 input 可能還沒 render，時序差會抓不到 input）。
    click_tab_by_text(page, "產品知識問答")
    page.wait_for_function(WAIT_VISIBLE_INPUT_JS, timeout=12000)

    # 3. 取得可見問題輸入框；找不到代表分頁狀態異常，往上拋由呼叫端處理。
    question_input = get_question_input(page)
    if question_input is None:
        raise RuntimeError("找不到產品知識問答分頁的可見輸入框（placeholder 不符）")

    # 4. 送出前先記錄目前答案作為新答案判定基準（fresh page 自然為空字串）。
    prev_answer = page.evaluate(EXTRACT_ANSWER_JS, ANSWER_MARKER)

    # 5. 用 Playwright 原生 fill() 注入問題：送真實 input 事件，Streamlit 才會把值
    #    commit 到 server-side session_state。不可按 Enter（會 rerun 把送出鈕洗掉）。
    question_input.fill(question)
    time.sleep(COMMIT_WAIT)  # 給 websocket commit 一點時間

    # 6. 點送出（用重試版：fill 後按鈕可能短暫重渲染被 detach）
    if not click_button_retry(page, "送出問題"):
        raise RuntimeError("重試多次仍找不到可見的「送出問題」按鈕")

    # 7. 輪詢擷取新答案（補點 tab 的時機已移到 extract_answer 輪詢途中處理）
    return extract_answer(page, prev_answer)


def open_browser(playwright, headed, bot_url):
    """啟動 Chromium、開頁、等 Streamlit ready，回傳 (browser, page)。"""
    browser = playwright.chromium.launch(headless=not headed)
    page = browser.new_page()
    page.goto(bot_url, timeout=60000)
    wait_streamlit_ready(page)
    return browser, page


def process_question(page, row_idx, question, bot_url):
    """處理單題並回傳答案結果字串。

    回傳值：答案文字、TIMEOUT_MARK，或拋例外（後端掛掉時）。
    後端例外是不可恢復的整批終止訊號，往上拋。
    """
    exc_text = check_backend_exception(page)
    if exc_text:
        raise BackendError(f"偵測到後端例外：\n{exc_text[:500]}")
    answer = ask_one_question(page, question, bot_url)
    if answer is None:
        print(f"  [TIMEOUT] Row {row_idx}：{question[:30]} … 等 "
              f"{ANSWER_POLL_TIMEOUT}s 無新答案", file=sys.stderr)
        return TIMEOUT_MARK
    return answer


def save_or_report(wb, excel_path, row_idx, question):
    """原子寫存檔；Excel 被鎖定時給出明確提示並中止整批。

    先 save 到同目錄暫存檔，成功後 os.replace 原子 rename 覆蓋原檔。避免直接覆寫
    時若寫到一半被硬中斷，整個 .xlsx 損毀丟失全部已答進度（reviewer 預警）。
    """
    tmp_path = excel_path.with_suffix(excel_path.suffix + ".tmp")
    try:
        wb.save(tmp_path)
        # 同分割區 rename 為原子操作：要嘛舊檔、要嘛新檔，不會出現半寫壞檔。
        os.replace(tmp_path, excel_path)
    except PermissionError:
        print(
            f"錯誤：無法寫入 Excel（Row {row_idx}：{question[:30]}）。\n"
            f"檔案可能正被 Excel 開啟並鎖定，請關閉後重跑。路徑：{excel_path}",
            file=sys.stderr,
        )
        # 清掉殘留暫存檔，避免下次 os.replace 撞到舊 tmp。
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        raise


def run_batch(ws, wb, excel_path, page, pending, dry_run):
    """逐題處理待答清單，回傳 (success, timeout) 計數。

    冪等保證：每題即存。錯誤防禦：單題瀏覽器崩潰重 goto 續跑、
    連續多題 TIMEOUT 視為後端掛掉而中止整批。
    """
    success, timeout, consecutive_timeout = 0, 0, 0
    bot_url = page.url
    for row_idx, question in pending:
        try:
            result = process_question(page, row_idx, question, bot_url)
        except PlaywrightError as exc:
            # 瀏覽器或頁面崩潰：重 goto 後續跑，不丟已存進度、不從頭。
            print(f"  [PAGE-CRASH] Row {row_idx}：{exc}，重新載入頁面後續跑",
                  file=sys.stderr)
            page.goto(bot_url, timeout=60000)
            wait_streamlit_ready(page)
            result = TIMEOUT_MARK
        except RuntimeError as exc:
            # 渲染類暫態錯誤（reload 後 tab/input/送出鈕尚未 render）：降級為單題失敗
            # 續跑，不殺整批（這正是 Row 262 整批中止的原因）。下一題 ask_one_question
            # 開頭會重 goto 開全新 session，不需在此額外 reload。BackendError 不在此
            # 捕捉，會穿透中止整批（後端真壞掉硬跑無意義）。
            print(f"  [RENDER-ERR] Row {row_idx}：{exc}，標記逾時續跑下一題",
                  file=sys.stderr)
            result = TIMEOUT_MARK

        if result == TIMEOUT_MARK:
            timeout += 1
            consecutive_timeout += 1
        else:
            success += 1
            consecutive_timeout = 0

        if dry_run:
            print(f"Row {row_idx} | {question[:40]} | {result[:80]}")
        else:
            ws.cell(row=row_idx, column=ANSWER_COL).value = result
            save_or_report(wb, excel_path, row_idx, question)

        if consecutive_timeout >= MAX_CONSECUTIVE_TIMEOUT:
            print(
                f"錯誤：連續 {MAX_CONSECUTIVE_TIMEOUT} 題 TIMEOUT，"
                "後端可能又進入冷啟動或掛了，中止整批。",
                file=sys.stderr,
            )
            raise SystemExit(2)

        time.sleep(THROTTLE_BETWEEN_QUESTIONS)
    return success, timeout


def parse_args():
    """解析 CLI 參數。"""
    parser = argparse.ArgumentParser(description="客服題庫批次自動化腳本")
    parser.add_argument("--dry-run", action="store_true",
                        help="只處理前 3 個待答題、印出結果、不寫不存 Excel")
    parser.add_argument("--limit", type=int, default=None,
                        help="最多處理 N 題（預設無上限）")
    parser.add_argument("--headed", action="store_true",
                        help="顯示瀏覽器視窗（預設 headless）")
    parser.add_argument("--bot-url", default=DEFAULT_BOT_URL,
                        help=f"Bot 網址（預設 {DEFAULT_BOT_URL}）")
    return parser.parse_args()


def main():
    """主流程：載入 Excel → 偵測待答 → 開瀏覽器 → 批次處理 → 回報。"""
    args = parse_args()
    excel_path = get_excel_path()
    if not excel_path.exists():
        print(f"錯誤：找不到題庫檔：{excel_path}", file=sys.stderr)
        sys.exit(1)

    # dry-run 只看前 3 題；--limit 另行限制。
    effective_limit = 3 if args.dry_run else args.limit

    try:
        wb = load_workbook(excel_path)
    except PermissionError:
        print(f"錯誤：Excel 被鎖定，請關閉後重跑：{excel_path}", file=sys.stderr)
        sys.exit(1)
    if SHEET_NAME not in wb.sheetnames:
        print(f"錯誤：工作表「{SHEET_NAME}」不存在，現有：{wb.sheetnames}",
              file=sys.stderr)
        sys.exit(1)
    ws = wb[SHEET_NAME]

    pending = find_resume_rows(ws, limit=effective_limit)
    if not pending:
        print("沒有待處理的題目（col8 皆已填）。")
        return
    print(f"待處理 {len(pending)} 題（dry_run={args.dry_run}），"
          f"起點 Row {pending[0][0]}")

    with sync_playwright() as playwright:
        browser, page = open_browser(playwright, args.headed, args.bot_url)
        try:
            success, timeout = run_batch(
                ws, wb, excel_path, page, pending, args.dry_run)
        finally:
            browser.close()

    print(f"完成：成功 {success} / TIMEOUT {timeout}")


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise  # 保留明確的 exit code（如連續 TIMEOUT 的 exit 2）
    except Exception as exc:  # noqa: BLE001 - 頂層守門，印環境資訊後非 0 退出
        print(
            f"錯誤：{exc}\n"
            f"環境：Python {sys.version.split()[0]}, platform={sys.platform}",
            file=sys.stderr,
        )
        sys.exit(1)
